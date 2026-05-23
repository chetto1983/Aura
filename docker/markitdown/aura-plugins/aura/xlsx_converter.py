"""
Per-sheet XLSX converter for Aura's wiki ingestion.

Each sheet becomes one '## Sheet: <name>' section with a clean markdown
table. Empty rows are skipped, the first non-empty row is promoted to
header, and fully-empty columns are dropped. No per-row '### Row N'
headings.

History: an earlier revision of this converter emitted '### Row N' for
every spreadsheet row to feed Aura's heading-subnode indexer. Paired
with the structured ingest splitter (internal/storage/sources/ingest/
structured.go) that turned every '###' into a separate wiki page, the
result was one wiki page per spreadsheet row — 876 single-row pages
from one 5-sheet workbook in 2026-05-23. Now sheets become H2 anchors
inside a single source page; the subnode indexer picks them up as
in-page subnodes without exploding the page count.

Cell value normalisation:
  None         -> '' (empty cell — never the literal string '(empty)')
  datetime     -> ISO 8601
  everything else -> str(), no precision loss for int/float
"""

import io
import warnings
from datetime import datetime
from typing import Any, BinaryIO

import openpyxl
from markitdown import DocumentConverter, DocumentConverterResult, StreamInfo

_XLSX_EXTENSIONS = {".xlsx"}
_XLSX_MIME_PREFIXES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)


def _cell_str(v: Any) -> str:
    """Normalise a cell value to its markdown representation.

    Returns the empty string for None — markdown tables render empty
    cells as blank, which is what we want. The previous '(empty)'
    literal multiplied across thousands of phantom rows accounted for
    most of the noise seen on Programma base.xlsx.
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v)


def _is_empty_cell(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v == "")


def _row_is_empty(row: tuple) -> bool:
    return all(_is_empty_cell(v) for v in row)


def _yaml_scalar(s: str) -> str:
    """Always-quoted YAML scalar; avoids parsing ambiguity for arbitrary strings."""
    safe = str(s).replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")
    return f'"{safe}"'


def _escape_cell(s: str) -> str:
    """Escape characters that would break a markdown table row."""
    return s.replace("|", "\\|").replace("\n", " ").replace("\r", " ")


def _prepare_sheet(rows_iter) -> tuple[list[str], list[list[str]]]:
    """Read a sheet's rows, drop empty rows + columns, promote first row to header.

    Returns (header, body). header is empty when the sheet contains no
    meaningful data (all rows empty, or no columns survive the prune).
    """
    rows: list[list] = []
    for row in rows_iter:
        if _row_is_empty(row):
            continue
        rows.append(list(row))

    if not rows:
        return [], []

    width = max(len(r) for r in rows)
    normalised: list[list] = [r + [None] * (width - len(r)) for r in rows]

    keep_cols: list[int] = []
    for ci in range(width):
        if any(not _is_empty_cell(normalised[ri][ci]) for ri in range(len(normalised))):
            keep_cols.append(ci)

    if not keep_cols:
        return [], []

    pruned = [[normalised[ri][ci] for ci in keep_cols] for ri in range(len(normalised))]

    header = [_cell_str(c) for c in pruned[0]]
    if all(h == "" for h in header):
        header = [f"col_{i + 1}" for i in range(len(header))]

    body = [[_cell_str(c) for c in r] for r in pruned[1:]]
    return header, body


class AuraXlsxConverter(DocumentConverter):
    """
    Per-sheet XLSX converter. Output shape:
      ---
      workbook_title: "products.xlsx"
      sheets:
        - "Products"
      non_empty_row_counts:
        "Products": 2
      author: "alice"
      ---

      ## Sheet: Products

      | Name | Price | Category |
      | --- | --- | --- |
      | Mozzarella | 2.50 | Dairy |
      | Prosciutto | 8.99 | Meat |
    """

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> bool:
        ext = (getattr(stream_info, "extension", "") or "").lower()
        mime = (getattr(stream_info, "mimetype", "") or "").lower()
        return ext in _XLSX_EXTENSIONS or any(
            mime.startswith(p) for p in _XLSX_MIME_PREFIXES
        )

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        data = file_stream.read()

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(
                io.BytesIO(data), data_only=True, read_only=True
            )

        try:
            creator = (wb.properties.creator or "").strip()
            author = creator if creator else "unknown"
        except Exception:
            author = "unknown"

        sheet_names = list(wb.sheetnames)
        sheet_blocks: list[tuple[str, list[str], list[list[str]]]] = []
        non_empty_counts: dict[str, int] = {}

        for sname in sheet_names:
            ws = wb[sname]
            header, body = _prepare_sheet(ws.iter_rows(values_only=True))
            non_empty_counts[sname] = len(body)
            sheet_blocks.append((sname, header, body))

        wb.close()

        filename = (getattr(stream_info, "filename", None) or "").strip()
        title = filename or "workbook"

        lines: list[str] = ["---"]
        lines.append(f"workbook_title: {_yaml_scalar(title)}")
        lines.append("sheets:")
        for s in sheet_names:
            lines.append(f"  - {_yaml_scalar(s)}")
        lines.append("non_empty_row_counts:")
        for s in sheet_names:
            lines.append(f"  {_yaml_scalar(s)}: {non_empty_counts.get(s, 0)}")
        lines.append(f"author: {_yaml_scalar(author)}")
        lines.append("---")
        lines.append("")

        for sname, header, body in sheet_blocks:
            lines.append(f"## Sheet: {sname}")
            lines.append("")
            if not header:
                lines.append("_(empty sheet)_")
                lines.append("")
                continue

            lines.append(
                "| " + " | ".join(_escape_cell(h) for h in header) + " |"
            )
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")
            for row in body:
                lines.append(
                    "| " + " | ".join(_escape_cell(c) for c in row) + " |"
                )
            lines.append("")

        return DocumentConverterResult(markdown="\n".join(lines).rstrip())
