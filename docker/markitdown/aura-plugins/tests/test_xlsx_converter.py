"""
Unit tests for AuraXlsxConverter.

These tests pin the per-sheet shape: one '## Sheet:' section per
worksheet containing a clean markdown table, with empty rows skipped,
empty columns dropped, and the first non-empty row promoted to header.

There must NEVER be any '### Row N' headings in the output — that was
the previous shape and it produced one wiki page per spreadsheet row
when fed through the structured ingest splitter.

Run with:  python -m pytest docker/markitdown/aura-plugins/tests/
"""

import io
import os
import sys

# Allow running from repo root without installing the package
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
import pytest
from markitdown import StreamInfo

from aura.xlsx_converter import AuraXlsxConverter


def _make_xlsx(sheets: dict) -> bytes:
    """Build an in-memory xlsx.

    sheets = {"SheetName": [header_row, data_row, ...]}
    """
    wb = openpyxl.Workbook()
    first = True
    for sname, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sname
            first = False
        else:
            ws = wb.create_sheet(sname)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stream_info(filename: str = "test.xlsx") -> StreamInfo:
    return StreamInfo(
        extension=".xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


# ---------------------------------------------------------------------------
# accepts() tests
# ---------------------------------------------------------------------------


def test_accepts_xlsx_extension():
    c = AuraXlsxConverter()
    assert c.accepts(io.BytesIO(b""), _stream_info())


def test_accepts_xlsx_mime():
    c = AuraXlsxConverter()
    si = StreamInfo(
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert c.accepts(io.BytesIO(b""), si)


def test_rejects_docx():
    c = AuraXlsxConverter()
    si = StreamInfo(
        extension=".docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )
    assert not c.accepts(io.BytesIO(b""), si)


# ---------------------------------------------------------------------------
# convert() shape tests
# ---------------------------------------------------------------------------


def test_convert_emits_sheet_section_and_markdown_table():
    data = _make_xlsx(
        {
            "Products": [
                ["Name", "Price", "Category"],
                ["Mozzarella", 2.50, "Dairy"],
                ["Prosciutto", 8.99, "Meat"],
            ]
        }
    )
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info("products.xlsx"))
    md = result.markdown

    assert md.startswith("---"), f"expected frontmatter, got: {md[:60]}"
    assert "workbook_title:" in md
    assert "## Sheet: Products" in md

    # Markdown table header + separator + body rows
    assert "| Name | Price | Category |" in md
    assert "| --- | --- | --- |" in md
    assert "| Mozzarella | 2.5 | Dairy |" in md
    assert "| Prosciutto | 8.99 | Meat |" in md


def test_convert_never_emits_row_headings():
    """The '### Row N' shape MUST stay dead.

    A converter that emits '### Row N' makes the structured ingest
    splitter materialise one wiki page per row, which polluted the
    graph with 876 single-row nodes from one workbook in 2026-05-23.
    """
    data = _make_xlsx(
        {
            "Sheet1": [
                ["A", "B"],
                [1, 2],
                [3, 4],
                [5, 6],
            ]
        }
    )
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info())
    assert "### Row" not in result.markdown
    assert "### " not in result.markdown


def test_convert_skips_fully_empty_rows():
    data = _make_xlsx(
        {
            "Sheet1": [
                ["A", "B"],
                [None, None],          # ← empty row, must be dropped
                ["x", "y"],
                [None, None],          # ← empty row, must be dropped
                ["w", "z"],
            ]
        }
    )
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info())
    md = result.markdown

    # 2 body rows after header, no empty stubs
    assert "| x | y |" in md
    assert "| w | z |" in md
    assert "non_empty_row_counts:" in md
    assert '"Sheet1": 2' in md or "Sheet1: 2" in md


def test_convert_drops_fully_empty_columns():
    data = _make_xlsx(
        {
            "Sheet1": [
                ["A", None, "B", None],
                ["a1", None, "b1", None],
                ["a2", None, "b2", None],
            ]
        }
    )
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info())
    md = result.markdown

    # 4 columns reduced to 2 (only A + B kept)
    assert "| A | B |" in md
    assert "| --- | --- |" in md
    assert "| a1 | b1 |" in md
    # The dropped empty columns must not appear as '|  | ' between header cells
    assert "| A |  | B |" not in md


def test_convert_empty_cell_renders_as_blank_not_literal_empty():
    """None must render as an empty markdown cell, NOT the literal '(empty)'.

    Regression guard for the 2026-05-23 catastrophe where each None cell
    on Programma base.xlsx became the literal text '(empty)' multiplied
    across 20 columns × 191 rows per sheet.
    """
    data = _make_xlsx(
        {
            "Sheet1": [
                ["Name", "Note"],
                ["Alice", None],
                ["Bob", "see SLA"],
            ]
        }
    )
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info())
    md = result.markdown
    assert "(empty)" not in md, f"converter emitted literal (empty):\n{md}"
    # Alice's note cell renders as an empty markdown cell
    assert "| Alice |  |" in md
    assert "| Bob | see SLA |" in md


def test_convert_int_cell_no_precision_loss():
    data = _make_xlsx({"Sheet1": [["ID"], [42]]})
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info())
    assert "| 42 |" in result.markdown


def test_convert_frontmatter_lists_sheets_and_counts():
    data = _make_xlsx(
        {
            "Data": [
                ["Col1", "Col2"],
                ["a", "b"],
                ["c", "d"],
                ["e", "f"],
            ]
        }
    )
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info("data.xlsx"))
    md = result.markdown
    assert "sheets:" in md
    assert '  - "Data"' in md
    # 3 body rows after header
    assert '"Data": 3' in md or "Data: 3" in md


def test_convert_multi_sheet():
    data = _make_xlsx(
        {
            "Alpha": [["X"], [1], [2]],
            "Beta": [["Y"], [10]],
        }
    )
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info("multi.xlsx"))
    md = result.markdown
    assert "## Sheet: Alpha" in md
    assert "## Sheet: Beta" in md
    # No row headings, just markdown tables
    assert "### Row" not in md
    assert "| 1 |" in md
    assert "| 10 |" in md


def test_convert_empty_sheet_no_crash():
    data = _make_xlsx({"Empty": []})
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info())
    md = result.markdown
    assert "## Sheet: Empty" in md
    assert "_(empty sheet)_" in md
    # Frontmatter still present
    assert "non_empty_row_counts:" in md
    assert '"Empty": 0' in md or "Empty: 0" in md


def test_convert_escapes_pipe_in_cell():
    data = _make_xlsx(
        {
            "Sheet1": [
                ["Tag", "Expr"],
                ["A|B", "x | y"],
            ]
        }
    )
    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info())
    md = result.markdown
    # Pipe in cell must be escaped so the table parses
    assert r"| A\|B | x \| y |" in md


def test_convert_first_row_with_padding_still_promoted_as_header():
    """First non-empty row is the header even when it carries whitespace
    padding cells. Documents that whitespace is preserved (we only
    fabricate generic 'col_N' names when the header is entirely empty
    strings after _cell_str normalisation)."""
    # openpyxl.Workbook().save(buf) leaves the buffer cursor at EOF;
    # use _make_xlsx so we get back bytes ready for BytesIO consumers.
    data = _make_xlsx(
        {
            "Sheet1": [
                ["Col A", " ", "Col C"],
                ["a", "x", "c"],
                ["d", "y", "f"],
            ]
        }
    )

    c = AuraXlsxConverter()
    result = c.convert(io.BytesIO(data), _stream_info())
    md = result.markdown
    assert "## Sheet: Sheet1" in md
    # Header preserved including the whitespace cell
    assert "| Col A |   | Col C |" in md
    assert "| a | x | c |" in md
    assert "| d | y | f |" in md


# ---------------------------------------------------------------------------
# Integration: use bench fixture if present
# ---------------------------------------------------------------------------


def test_convert_bench_fixture_if_present():
    fixture = os.path.join(
        os.path.dirname(__file__),
        "../../../../docs/quality-bench/fixtures/tika-testEXCEL.xlsx",
    )
    if not os.path.exists(fixture):
        pytest.skip("bench fixture not present")

    with open(fixture, "rb") as f:
        data = f.read()

    c = AuraXlsxConverter()
    result = c.convert(
        io.BytesIO(data),
        StreamInfo(extension=".xlsx", filename="tika-testEXCEL.xlsx"),
    )
    md = result.markdown
    assert "## Sheet:" in md, "no sheet sections in bench fixture output"
    assert "### Row" not in md, "regression: row-per-page shape leaked back in"
    assert "non_empty_row_counts:" in md
