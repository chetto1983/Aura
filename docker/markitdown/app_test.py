import os
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook

import app


class XlsxExtractionTests(unittest.TestCase):
    def test_internal_empty_cell_keeps_stable_coordinates(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet["A1"] = "left"
        sheet["C1"] = "right"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
            path = handle.name
        try:
            workbook.save(path)
            payload = app._extract_xlsx(path)
        finally:
            workbook.close()
            os.remove(path)

        self.assertEqual(payload["stats"]["sheets"], 1)
        text = "\n".join(chunk["text"] for chunk in payload["chunks"])
        self.assertIn("A1: left", text)
        self.assertIn("C1: right", text)

    def test_coordinate_does_not_depend_on_cell_attributes(self) -> None:
        class Cell:
            def __init__(self, value: object) -> None:
                self.value = value

        class Sheet:
            title = "Data"

            @staticmethod
            def iter_rows():
                return iter([[Cell("left"), Cell(None), Cell("right")]])

        class WorkbookFixture:
            worksheets = [Sheet()]

            @staticmethod
            def close() -> None:
                return None

        with patch("openpyxl.load_workbook", return_value=WorkbookFixture()):
            payload = app._extract_xlsx("fixture.xlsx")
        text = "\n".join(chunk["text"] for chunk in payload["chunks"])
        self.assertIn("A1: left", text)
        self.assertIn("C1: right", text)


if __name__ == "__main__":
    unittest.main()
