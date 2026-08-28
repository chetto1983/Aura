"""Build the two hash-pinned fixtures that close the document-agent oracle gap.

The workbook is deliberately source data only: the requested filtered extension total
does not appear in any cell, formula, card or indexed chunk as a precomputed answer. The
agent must open the original file and calculate it. The PDF contains one raster image and
no text operators, so a parser-only pipeline cannot accidentally satisfy the OCR case.
"""

import argparse
import datetime as dt
import hashlib
import re
import tempfile
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageDraw, ImageFont
from pypdf import PdfReader
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas


FIXED_ZIP_TIME = (2026, 1, 1, 0, 0, 0)
FIXED_DOC_TIME = dt.datetime(2026, 1, 1, tzinfo=dt.timezone.utc)
WORKBOOK_NAME = "movimenti_analitici_5000.xlsx"
SCAN_NAME = "verbale_scansionato_italiano.pdf"


def movement_rows():
    centers = ("ALFA", "BETA", "GAMMA")
    for index in range(1, 5001):
        yield (
            f"MOV-{index:05d}",
            centers[(index - 1) % len(centers)],
            "VALIDO" if index % 5 else "ANNULLATO",
            (index % 7) + 1,
            ((index % 19) + 1) * 10,
        )


def write_workbook(path: Path) -> int:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimenti"
    headers = ("ID Movimento", "Centro", "Stato", "Quantita", "PrezzoUnitario")
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    expected = 0
    for row in movement_rows():
        sheet.append(row)
        if row[1] == "ALFA" and row[2] == "VALIDO":
            expected += row[3] * row[4]
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    for column, width in zip("ABCDE", (18, 12, 14, 12, 20), strict=True):
        sheet.column_dimensions[column].width = width
    workbook.properties.creator = "Aura synthetic document oracle"
    workbook.properties.created = FIXED_DOC_TIME
    workbook.properties.modified = FIXED_DOC_TIME

    with tempfile.TemporaryDirectory() as scratch:
        raw = Path(scratch) / path.name
        workbook.save(raw)
        with zipfile.ZipFile(raw, "r") as source, zipfile.ZipFile(
            path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as target:
            for info in sorted(source.infolist(), key=lambda item: item.filename):
                normalized = zipfile.ZipInfo(info.filename, FIXED_ZIP_TIME)
                normalized.compress_type = zipfile.ZIP_DEFLATED
                normalized.create_system = 3
                normalized.external_attr = info.external_attr
                content = source.read(info.filename)
                if info.filename == "docProps/core.xml":
                    content = re.sub(
                        rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                        rb"\g<1>2026-01-01T00:00:00Z\g<2>",
                        content,
                    )
                target.writestr(normalized, content)
    return expected


def write_scan(path: Path) -> None:
    image = Image.new("RGB", (2480, 3508), "white")
    draw = ImageDraw.Draw(image)
    font_path = "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"
    title = ImageFont.truetype(font_path, 92)
    body = ImageFont.truetype(font_path, 74)
    draw.text((180, 240), "VERBALE SCANSIONATO AURA", fill="black", font=title)
    for row, text in enumerate(
        (
            "CODICE PRATICA: ITALIA-7391",
            "IMPORTO APPROVATO: 48270 EURO",
            "RESPONSABILE: GIULIA BIANCHI",
        ),
        start=1,
    ):
        draw.text((180, 520 + row * 190), text, fill="black", font=body)
    pdf = canvas.Canvas(str(path), pagesize=A4, pageCompression=0, invariant=1)
    pdf.drawImage(ImageReader(image), 0, 0, width=A4[0], height=A4[1])
    pdf.showPage()
    pdf.save()
    extracted = "".join(page.extract_text() or "" for page in PdfReader(path).pages)
    if extracted.strip():
        raise RuntimeError(f"{path.name} unexpectedly contains a text layer: {extracted!r}")


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()
    args.out.mkdir(parents=True, exist_ok=True)
    workbook = args.out / WORKBOOK_NAME
    scan = args.out / SCAN_NAME
    expected = write_workbook(workbook)
    write_scan(scan)
    print(f"{WORKBOOK_NAME}: sha256={digest(workbook)} expected_filtered_total={expected}")
    print(f"{SCAN_NAME}: sha256={digest(scan)} text_layer=empty")


if __name__ == "__main__":
    main()
