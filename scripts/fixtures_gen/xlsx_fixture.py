"""The spreadsheet fixture, built with openpyxl as the xlsx skill prescribes.

It is not a generic table: it reproduces the exact shape `internal/documents/filecard`
has to survive, because that is what a real operator file looks like and what the
old pipeline could not open at all.

  row 1  a merged banner        -- junk above the data
  row 2  blank                  -- more junk
  row 3  the REAL header        -- what filecard must find
  row 4+ data, numbers as NUMBERS

filecard's own TestHeaderIsFoundBelowABannerRow covers this shape; a fixture whose
header sits on row 1 would pass a pipeline that cannot handle any real file.
"""

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

CANARY_CODE = "A9A26924"

wb = Workbook()
ws = wb.active
ws.title = "Ordine"

ws.merge_cells("A1:C1")
ws["A1"] = "ORDINE DI PROVA - CORPUS SINTETICO AURA"
ws["A1"].font = Font(name="Arial", size=14, bold=True)

header = ["Codice", "Quantita", "Descrizione"]
for col, value in enumerate(header, start=1):
    cell = ws.cell(row=3, column=col, value=value)
    cell.font = Font(name="Arial", bold=True)

rows = [
    ("28908", 1, "Interruttore automatico di prova"),
    ("28958", 3, "Copriviti di prova per interruttore"),
    (CANARY_CODE, 11, "Contatto ausiliario aperto-chiuso di prova"),
    ("LV429630", 2, "Blocco differenziale di prova"),
    ("A9F74216", 7, "Interruttore modulare di prova"),
]
for r, (code, qty, desc) in enumerate(rows, start=4):
    ws.cell(row=r, column=1, value=code)
    ws.cell(row=r, column=2, value=qty)      # int, not str -- the ETL path needs a number
    ws.cell(row=r, column=3, value=desc)

for col, width in enumerate((16, 12, 46), start=1):
    ws.column_dimensions[get_column_letter(col)].width = width

wb.save("/out/sample.xlsx")
total = sum(q for _, q, _ in rows)
print(f"sample.xlsx: {len(rows)} righe, header in riga 3 sotto un banner, totale quantita {total}")
