"""Benchmark: card-catalog routing vs chunk-flatten (markitdown) routing.

Question under test (the operator's "questione fondamentale"): given a query,
can we route to the RIGHT file(s) without opening every blob?

We compare two REPRESENTATIONS of the same corpus, searched with the SAME
BM25 engine (so any difference is representation, not algorithm):

  A. chunk_flatten  -- replicates docker/markitdown/app.py _extract_xlsx:
       every cell -> "{col}{row}: {value}", rows joined, chunked every 50 rows
       / 12k chars, whitespace collapsed. One BM25 doc per CHUNK.
  B. card_catalog   -- one compact "card" per FILE: filename + sheet names +
       distinct short text values (labels/headers/categoricals/entities,
       deduped) + numeric ranges. One BM25 doc per FILE.

Both representations include the filename (parity), so the comparison isolates
the effect of how CONTENT is represented, not a filename freebie.

Metrics: Recall@1, Recall@3, MRR of the ground-truth target file, overall and
by query type. Plus an "answerability" demonstration: for a computed-aggregate
query, the answer exists in NO chunk -- routing is necessary but not sufficient,
open+compute is the only path to the number.
"""

import math
import os
import re
from collections import defaultdict
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
CORPUS = os.path.join(HERE, "corpus")

MAX_CHUNK_CHARS = 12000
MAX_CHUNK_ROWS = 50

# ----------------------------------------------------------------------------
# tokenization
# ----------------------------------------------------------------------------
_ACCENTS = str.maketrans("àáâãäèéêëìíîïòóôõöùúûü", "aaaaaeeeeiiiiooooouuuu")


def tok(text: str):
    text = text.lower().translate(_ACCENTS)
    return [t for t in re.split(r"[^a-z0-9]+", text) if len(t) >= 2 or t.isdigit()]


# ----------------------------------------------------------------------------
# reading files
# ----------------------------------------------------------------------------
def read_xlsx_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []  # list of (sheet_title, [ [cellstr,...], ... ])
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c).strip() for c in row])
        out.append((ws.title, rows))
    wb.close()
    return out


def read_txt(path):
    with open(path, encoding="utf-8") as f:
        return f.read()


# ----------------------------------------------------------------------------
# representation A: chunk_flatten (markitdown parity)
# ----------------------------------------------------------------------------
def flatten_xlsx_chunks(path):
    from openpyxl.utils import get_column_letter
    chunks = []
    for sheet, rows in read_xlsx_rows(path):
        lines, n = [], 0
        for ri, row in enumerate(rows, start=1):
            cells = []
            for ci, val in enumerate(row, start=1):
                if val == "":
                    continue
                cells.append(f"{get_column_letter(ci)}{ri}: {val}")
            if not cells:
                continue
            lines.append(f"row {ri}: " + " | ".join(cells))
            n += 1
            joined = "\n".join(lines)
            if len(joined) >= MAX_CHUNK_CHARS or n >= MAX_CHUNK_ROWS:
                chunks.append(re.sub(r"\s+", " ", joined))
                lines, n = [], 0
        if lines:
            chunks.append(re.sub(r"\s+", " ", "\n".join(lines)))
    return chunks or [""]


def flatten_txt_chunks(path, size=800):
    text = re.sub(r"\s+", " ", read_txt(path))
    return [text[i:i + size] for i in range(0, len(text), size)] or [""]


# ----------------------------------------------------------------------------
# representation B: card_catalog (schema fingerprint per file)
# ----------------------------------------------------------------------------
INV_RE = re.compile(r"\b20\d\d[_/]?\d{3}\b")


def card_xlsx(path):
    parts = []
    labels = set()
    entities = set()
    nums = []
    sheets = []
    for sheet, rows in read_xlsx_rows(path):
        sheets.append(sheet)
        for row in rows:
            for val in row:
                if val == "":
                    continue
                # numeric -> feed range, not the value itself
                v = val.replace(",", "")
                try:
                    nums.append(float(v))
                    continue
                except ValueError:
                    pass
                if len(val) <= 40:
                    labels.add(val)
                for m in INV_RE.findall(val):
                    entities.add(m)
    parts.append("fogli: " + ", ".join(sheets))
    # dedup labels; cap so the card stays compact (the whole point)
    parts.append("campi e valori: " + " | ".join(sorted(labels)[:80]))
    if entities:
        parts.append("riferimenti fattura: " + " ".join(sorted(entities)))
    if nums:
        parts.append(f"importi da {min(nums):.2f} a {max(nums):.2f}")
    return "\n".join(parts)


def card_txt(path, words=60):
    text = read_txt(path)
    head = " ".join(text.split()[:words])
    ents = set(INV_RE.findall(text))
    card = head
    if ents:
        card += "\nriferimenti fattura: " + " ".join(sorted(ents))
    return card


# ----------------------------------------------------------------------------
# BM25 (Okapi)
# ----------------------------------------------------------------------------
class BM25:
    def __init__(self, docs, k1=1.5, b=0.75):
        # docs: list of (doc_id, token_list)
        self.k1, self.b = k1, b
        self.ids = [d[0] for d in docs]
        self.toks = [d[1] for d in docs]
        self.dl = [len(t) for t in self.toks]
        self.avgdl = (sum(self.dl) / len(self.dl)) if self.dl else 0.0
        self.tf = [defaultdict(int) for _ in docs]
        df = defaultdict(int)
        for i, t in enumerate(self.toks):
            seen = set()
            for w in t:
                self.tf[i][w] += 1
                if w not in seen:
                    df[w] += 1
                    seen.add(w)
        n = len(docs)
        self.idf = {w: math.log(1 + (n - c + 0.5) / (c + 0.5)) for w, c in df.items()}

    def score(self, q_tokens, i):
        s = 0.0
        dl, avgdl = self.dl[i], self.avgdl or 1.0
        for w in q_tokens:
            if w not in self.tf[i]:
                continue
            f = self.tf[i][w]
            idf = self.idf.get(w, 0.0)
            s += idf * (f * (self.k1 + 1)) / (f + self.k1 * (1 - self.b + self.b * dl / avgdl))
        return s

    def rank(self, query):
        q = tok(query)
        scored = [(self.ids[i], self.score(q, i)) for i in range(len(self.ids))]
        scored.sort(key=lambda x: x[1], reverse=True)
        return scored


# ----------------------------------------------------------------------------
# build both indexes
# ----------------------------------------------------------------------------
def build_indexes():
    files = sorted(os.listdir(CORPUS))
    flatten_docs = []   # (chunk_id, tokens) chunk_id = "file#idx"
    card_docs = []      # (file, tokens)
    for f in files:
        path = os.path.join(CORPUS, f)
        if f.endswith(".xlsx"):
            chunks = flatten_xlsx_chunks(path)
            card = card_xlsx(path)
        else:
            chunks = flatten_txt_chunks(path)
            card = card_txt(path)
        for j, ch in enumerate(chunks):
            flatten_docs.append((f"{f}#{j}", tok(f + " " + ch)))  # filename parity
        card_docs.append((f, tok(f + " " + card)))
    return files, BM25(flatten_docs), BM25(card_docs)


def rank_files_flatten(bm25, query):
    """Aggregate chunk ranking -> file ranking (best chunk per file)."""
    best = {}
    for chunk_id, sc in bm25.rank(query):
        f = chunk_id.split("#")[0]
        if f not in best or sc > best[f]:
            best[f] = sc
    ordered = sorted(best.items(), key=lambda x: x[1], reverse=True)
    return [f for f, _ in ordered]


def rank_files_card(bm25, query):
    return [f for f, _ in bm25.rank(query)]


# ----------------------------------------------------------------------------
# query set: (query, {acceptable target files}, type)
# ----------------------------------------------------------------------------
Q = [
    ("dettaglio della fattura 2026_007", {"fattura_2026_007.xlsx"}, "EXACT"),
    ("voglio vedere la fattura numero 011", {"fattura_2026_011.xlsx"}, "EXACT"),
    ("fattura di febbraio emessa a Bianchi", {"fattura_2026_003.xlsx", "contabilita_2026.xlsx"}, "EXACT"),
    ("quanto abbiamo fatturato in totale ad ACME nel 2026", {"contabilita_2026.xlsx"}, "AGG"),
    ("totale IVA del secondo trimestre", {"contabilita_2026.xlsx"}, "AGG"),
    ("qual e' il totale complessivo fatturato nell'anno", {"contabilita_2026.xlsx"}, "AGG"),
    ("qual e' la fattura piu alta emessa a Verdi Costruzioni", {"contabilita_2026.xlsx", "fattura_2026_009.xlsx"}, "AGG"),
    ("quali fatture non sono state pagate", {"contabilita_2026.xlsx", "scadenzario.xlsx"}, "FILTER"),
    ("tutte le fatture del cliente Gialli Import", {"contabilita_2026.xlsx", "fattura_2026_007.xlsx", "fattura_2026_011.xlsx"}, "FILTER"),
    ("quali clienti sono in ritardo con i pagamenti", {"scadenzario.xlsx", "note_pagamenti.txt"}, "AGING"),
    ("quanti giorni di ritardo ha Verdi Costruzioni", {"scadenzario.xlsx", "note_pagamenti.txt"}, "AGING"),
    ("perche Bianchi ha contestato una fattura", {"note_pagamenti.txt"}, "PROSE"),
    ("qual e' il problema con la fattura 2026_006", {"note_pagamenti.txt"}, "PROSE"),
    ("quali sono i termini di pagamento di Verdi Costruzioni", {"anagrafica_clienti.xlsx"}, "CROSSREF"),
    ("la partita IVA di Rossi SRL", {"anagrafica_clienti.xlsx"}, "CROSSREF"),
    ("cosa prevede il contratto con ACME", {"contratto_ACME_2026.txt"}, "PROSE"),
    ("qual e' il foro competente del contratto ACME", {"contratto_ACME_2026.txt"}, "PROSE"),
    ("fatture emesse a marzo 2026", {"contabilita_2026.xlsx", "fattura_2026_005.xlsx", "fattura_2026_006.xlsx", "fattura_2026_007.xlsx"}, "TEMPORAL"),
    ("riepilogo generale dell'anno", {"riepilogo_annuale.txt"}, "PROSE"),
    ("come sono archiviati i documenti amministrativi", {"istruzioni_archivio.txt"}, "PROSE"),
]


def rank_of_target(ranked_files, targets):
    for i, f in enumerate(ranked_files, start=1):
        if f in targets:
            return i
    return None


def evaluate(name, ranker):
    per_type = defaultdict(lambda: [0, 0, 0.0, 0])  # r1, r3, mrr_sum, count
    rows = []
    for query, targets, qtype in Q:
        ranked = ranker(query)
        r = rank_of_target(ranked, targets)
        r1 = 1 if r == 1 else 0
        r3 = 1 if (r is not None and r <= 3) else 0
        mrr = (1.0 / r) if r else 0.0
        rows.append((query, qtype, r, r1, r3, mrr, ranked[:3]))
        agg = per_type[qtype]
        agg[0] += r1; agg[1] += r3; agg[2] += mrr; agg[3] += 1
    tot = [0, 0, 0.0, 0]
    for agg in per_type.values():
        tot[0] += agg[0]; tot[1] += agg[1]; tot[2] += agg[2]; tot[3] += agg[3]
    return name, rows, per_type, tot


def print_report(results):
    for name, rows, per_type, tot in results:
        n = tot[3]
        print(f"\n===== {name} =====")
        print(f"  overall   R@1={tot[0]}/{n} ({tot[0]/n:.0%})   R@3={tot[1]}/{n} ({tot[1]/n:.0%})   MRR={tot[2]/n:.3f}")
        print("  by type:")
        for qtype in sorted(per_type):
            a = per_type[qtype]
            print(f"    {qtype:9s} R@1={a[0]}/{a[3]}  R@3={a[1]}/{a[3]}  MRR={a[2]/a[3]:.3f}")


def print_side_by_side(res_flat, res_card):
    print("\n===== per-query (rank of target; lower is better; - = miss) =====")
    print(f"  {'type':9s} {'flat':>4s} {'card':>4s}  query")
    _, rf, _, _ = res_flat
    _, rc, _, _ = res_card
    for (q, t, rk_f, *_), (_, _, rk_c, *_) in zip(rf, rc):
        sf = str(rk_f) if rk_f else "-"
        sc = str(rk_c) if rk_c else "-"
        flag = "  <-- card better" if (rk_c or 99) < (rk_f or 99) else ("  <-- flat better" if (rk_f or 99) < (rk_c or 99) else "")
        print(f"  {t:9s} {sf:>4s} {sc:>4s}  {q}{flag}")


# ----------------------------------------------------------------------------
# answerability demonstration
# ----------------------------------------------------------------------------
def answerability_demo(flat_bm25):
    print("\n===== answerability: routing is necessary but NOT sufficient =====")
    # answer computed over the ledger; NOT stored as any single cell
    rows = read_xlsx_rows(os.path.join(CORPUS, "contabilita_2026.xlsx"))
    registro = dict(rows)["registro"]
    header = registro[0]
    ci = {h: i for i, h in enumerate(header)}
    unpaid = [r for r in registro[1:] if r and r[ci["stato"]] in ("non pagata", "scaduta")]
    totals = [float(r[ci["totale"]]) for r in unpaid]
    mean_unpaid = sum(totals) / len(totals)
    query = "importo medio delle fatture non pagate o scadute"
    print(f"  query: {query!r}")
    print(f"  open+compute answer: media = {mean_unpaid:.2f} EUR  (su {len(totals)} fatture)")
    # is that number present as a token in ANY flatten chunk?
    token = str(round(mean_unpaid, 2)).translate(_ACCENTS)
    present = any(token in " ".join(doc) for doc in flat_bm25.toks)
    print(f"  the number {mean_unpaid:.2f} appears in a flattened chunk? {present}")
    print("  -> even a perfect router only finds the FILE; the NUMBER exists")
    print("     in no chunk. Only open+compute (soffice/openpyxl) produces it.")


def main():
    files, flat_bm25, card_bm25 = build_indexes()
    n_flat_docs = len(flat_bm25.ids)
    print(f"corpus: {len(files)} files")
    print(f"  chunk_flatten index: {n_flat_docs} docs (chunks)")
    print(f"  card_catalog  index: {len(card_bm25.ids)} docs (one per file)")

    res_flat = evaluate("chunk_flatten (markitdown baseline)", lambda q: rank_files_flatten(flat_bm25, q))
    res_card = evaluate("card_catalog (schema fingerprint)", lambda q: rank_files_card(card_bm25, q))
    print_report([res_flat, res_card])
    print_side_by_side(res_flat, res_card)
    answerability_demo(flat_bm25)


if __name__ == "__main__":
    main()
