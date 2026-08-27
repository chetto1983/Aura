"""Generate a small, realistic, INTERLINKED Italian-accounting corpus.

21 files that reference each other:
  - 14 invoice workbooks (fattura_2026_NNN.xlsx)
  - 1 master ledger (contabilita_2026.xlsx)   <- links every invoice
  - 1 client registry (anagrafica_clienti.xlsx)
  - 1 aging/due-dates sheet (scadenzario.xlsx) <- links invoices
  - 4 prose files (.txt) referencing specific invoices/clients

The data is deterministic (no RNG) and internally consistent: the ledger totals
are the exact sum of the invoices, the aging sheet references real invoice ids,
and the prose notes cite real invoice numbers. That consistency is what makes
the routing question real -- the answer to an aggregation query lives in ONE
file (the ledger), computed from the others.
"""

import argparse
import datetime as dt
import hashlib
import os
import re
import zipfile
from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUT = os.path.abspath(os.path.join(
    HERE, "..", "..", "scripts", "fixtures", "document_retrieval_eval", "corpus"
))
OUT = DEFAULT_OUT
FIXED_DOCUMENT_TIME = dt.datetime(2026, 1, 1, tzinfo=dt.timezone.utc)
FIXED_ZIP_TIME = (2026, 1, 1, 0, 0, 0)

VAT = 0.22

# (name, piva, terms_days)
CLIENTS = [
    ("ACME SpA", "IT01234567890", 30),
    ("Rossi SRL", "IT02345678901", 60),
    ("Bianchi & Figli", "IT03456789012", 30),
    ("Verdi Costruzioni", "IT04567890123", 90),
    ("Gialli Import", "IT05678901234", 60),
]
CLIENT_NAMES = [c[0] for c in CLIENTS]

# One row per invoice. (num, client_idx, date, imponibile, stato)
# Dates spread Jan-Jun 2026 (Q1 = 01-03, Q2 = 04-06).
INVOICES = [
    (1,  0, "2026-01-12", 12000.0, "pagata"),
    (2,  1, "2026-01-28",  4500.0, "pagata"),
    (3,  2, "2026-02-05",  8200.0, "pagata"),
    (4,  0, "2026-02-18", 15750.0, "pagata"),
    (5,  3, "2026-03-03", 23000.0, "non pagata"),
    (6,  2, "2026-03-15",  6100.0, "contestata"),
    (7,  4, "2026-03-27",  9800.0, "pagata"),
    (8,  1, "2026-04-08",  5300.0, "pagata"),
    (9,  3, "2026-04-19", 31200.0, "scaduta"),
    (10, 0, "2026-04-30", 14200.0, "pagata"),
    (11, 4, "2026-05-11",  7600.0, "non pagata"),
    (12, 2, "2026-05-22",  8900.0, "pagata"),
    (13, 0, "2026-06-04", 16400.0, "non pagata"),
    (14, 1, "2026-06-16",  4950.0, "pagata"),
]


def inv_id(num: int) -> str:
    return f"2026_{num:03d}"


def money(x: float) -> float:
    return round(x, 2)


def derive(num, client_idx, date, imponibile, stato):
    iva = money(imponibile * VAT)
    totale = money(imponibile + iva)
    return {
        "id": inv_id(num),
        "cliente": CLIENTS[client_idx][0],
        "piva": CLIENTS[client_idx][1],
        "data": date,
        "imponibile": money(imponibile),
        "iva": iva,
        "totale": totale,
        "stato": stato,
    }


ROWS = [derive(*r) for r in INVOICES]


def save_workbook(wb, filename):
    """Write byte-for-byte stable XLSX fixtures, including ZIP metadata."""
    wb.properties.created = FIXED_DOCUMENT_TIME
    wb.properties.modified = FIXED_DOCUMENT_TIME
    target = os.path.join(OUT, filename)
    wb.save(target)

    normalized = target + ".normalized"
    with zipfile.ZipFile(target, "r") as source, zipfile.ZipFile(
        normalized, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as destination:
        for source_info in source.infolist():
            info = zipfile.ZipInfo(source_info.filename, FIXED_ZIP_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = source_info.external_attr
            payload = source.read(source_info.filename)
            # openpyxl overwrites modified with wall-clock time during save even
            # when Workbook.properties.modified was set immediately beforehand.
            if source_info.filename == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified\b[^>]*>).*?(</dcterms:modified>)",
                    rb"\g<1>2026-01-01T00:00:00Z\g<2>",
                    payload,
                )
            destination.writestr(info, payload)
    os.replace(normalized, target)


def write_invoice(inv):
    wb = Workbook()
    ws = wb.active
    ws.title = "fattura"
    ws.append(["Fattura n.", inv["id"]])
    ws.append(["Cliente", inv["cliente"]])
    ws.append(["Partita IVA", inv["piva"]])
    ws.append(["Data emissione", inv["data"]])
    ws.append(["Stato", inv["stato"]])
    ws.append([])
    ws.append(["Descrizione", "Quantita", "Prezzo unitario", "Importo"])
    # one synthetic line item that sums to imponibile
    ws.append(["Prestazione servizi", 1, inv["imponibile"], inv["imponibile"]])
    ws.append([])
    ws.append(["Imponibile", inv["imponibile"]])
    ws.append(["IVA 22%", inv["iva"]])
    ws.append(["Totale documento", inv["totale"]])
    save_workbook(wb, f"fattura_{inv['id']}.xlsx")


def write_ledger():
    wb = Workbook()
    reg = wb.active
    reg.title = "registro"
    reg.append(["n_fattura", "cliente", "data", "imponibile", "iva", "totale", "stato"])
    for r in ROWS:
        reg.append([r["id"], r["cliente"], r["data"], r["imponibile"], r["iva"], r["totale"], r["stato"]])

    rip = wb.create_sheet("riepilogo")
    rip.append(["cliente", "num_fatture", "imponibile_totale", "iva_totale", "totale_totale"])
    for name in CLIENT_NAMES:
        sub = [r for r in ROWS if r["cliente"] == name]
        rip.append([
            name, len(sub),
            money(sum(r["imponibile"] for r in sub)),
            money(sum(r["iva"] for r in sub)),
            money(sum(r["totale"] for r in sub)),
        ])
    rip.append([])
    rip.append(["trimestre", "imponibile", "iva", "totale"])
    for q, months in (("Q1", ("01", "02", "03")), ("Q2", ("04", "05", "06"))):
        sub = [r for r in ROWS if r["data"][5:7] in months]
        rip.append([
            q,
            money(sum(r["imponibile"] for r in sub)),
            money(sum(r["iva"] for r in sub)),
            money(sum(r["totale"] for r in sub)),
        ])
    save_workbook(wb, "contabilita_2026.xlsx")


def write_registry():
    wb = Workbook()
    ws = wb.active
    ws.title = "clienti"
    ws.append(["cliente", "partita_iva", "indirizzo", "email", "termini_pagamento_giorni"])
    addr = {
        "ACME SpA": "Via Roma 1, Milano",
        "Rossi SRL": "Corso Italia 22, Torino",
        "Bianchi & Figli": "Via Verdi 8, Bologna",
        "Verdi Costruzioni": "Viale Europa 100, Roma",
        "Gialli Import": "Via Mazzini 5, Genova",
    }
    for name, piva, terms in CLIENTS:
        slug = name.split()[0].lower()
        ws.append([name, piva, addr[name], f"info@{slug}.it", terms])
    save_workbook(wb, "anagrafica_clienti.xlsx")


def write_aging():
    wb = Workbook()
    ws = wb.active
    ws.title = "scadenzario"
    ws.append(["n_fattura", "cliente", "data_scadenza", "giorni_ritardo", "importo", "stato"])
    # only unpaid/overdue invoices appear in the aging sheet
    aging = {
        "2026_005": ("2026-04-02", 0),
        "2026_009": ("2026-07-18", 45),
        "2026_011": ("2026-07-10", 12),
        "2026_013": ("2026-07-04", 0),
    }
    for r in ROWS:
        if r["id"] in aging:
            scad, ritardo = aging[r["id"]]
            ws.append([r["id"], r["cliente"], scad, ritardo, r["totale"], r["stato"]])
    save_workbook(wb, "scadenzario.xlsx")


def write_txt(name, text):
    # Stable across Windows and Linux checkouts: Git must not silently change the
    # bytes after corpus.sha256 has been computed.
    with open(os.path.join(OUT, name), "w", encoding="utf-8", newline="\n") as f:
        f.write(text.strip() + "\n")


def write_prose():
    write_txt("note_pagamenti.txt", """
Note interne sui pagamenti - 2026

ACME SpA salda sempre puntualmente entro 30 giorni; nessuna criticita.
Bianchi & Figli ha contestato la fattura 2026_006 per un errore di quantita
sulla riga di prestazione: in attesa di nota di credito, importo sospeso.
Verdi Costruzioni risulta in ritardo di 45 giorni sulla fattura 2026_009,
sollecito inviato via PEC. La fattura 2026_005 e' ancora non pagata ma nei
termini contrattuali (90 giorni). Gialli Import: fattura 2026_011 non ancora
incassata. Rossi SRL nessun problema.
""")

    write_txt("contratto_ACME_2026.txt", """
Contratto di fornitura servizi - ACME SpA

Oggetto: prestazione di servizi di consulenza continuativa per l'anno 2026.
Corrispettivo: fatturazione mensile a consuntivo, IVA 22% esclusa.
Termini di pagamento: bonifico a 30 giorni data fattura.
Penali: interessi di mora ai sensi del D.Lgs 231/2002 in caso di ritardo.
Foro competente: Milano. Referente amministrativo: ufficio contabilita.
""")

    write_txt("riepilogo_annuale.txt", """
Riepilogo annuale 2026 (bozza)

Nel corso del 2026 sono state emesse 14 fatture verso 5 clienti principali.
Il cliente con il maggior volume e' ACME SpA. Alcune posizioni restano aperte
(fatture non pagate o scadute), monitorate nello scadenzario. La contabilita
completa e riconciliata e' mantenuta nel registro contabilita_2026.
""")

    write_txt("istruzioni_archivio.txt", """
Istruzioni archivio documenti amministrativi

Le fatture emesse sono archiviate singolarmente come file fattura_2026_NNN.
Il registro contabile contabilita_2026 contiene una riga per ogni fattura e i
riepiloghi per cliente e per trimestre. L'anagrafica clienti riporta partite
IVA, indirizzi e termini di pagamento. Lo scadenzario elenca le posizioni aperte.
""")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", default=DEFAULT_OUT, help="fixture output directory")
    args = parser.parse_args()
    global OUT
    OUT = os.path.abspath(args.out)
    os.makedirs(OUT, exist_ok=True)
    for inv in ROWS:
        write_invoice(inv)
    write_ledger()
    write_registry()
    write_aging()
    write_prose()
    files = sorted(os.listdir(OUT))
    checksum_path = os.path.join(os.path.dirname(OUT), "corpus.sha256")
    corpus_name = os.path.basename(OUT)
    with open(checksum_path, "w", encoding="ascii", newline="\n") as checksums:
        for filename in files:
            with open(os.path.join(OUT, filename), "rb") as fixture:
                digest = hashlib.sha256(fixture.read()).hexdigest()
            checksums.write(f"{digest}  {corpus_name}/{filename}\n")
    print(f"generated {len(files)} files in {OUT}")
    print(f"wrote checksums to {checksum_path}")
    for f in files:
        print("  ", f)


if __name__ == "__main__":
    main()
