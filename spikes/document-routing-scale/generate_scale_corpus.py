"""Generate an ENTERPRISE-SCALE corpus to stress the routing/retrieval design.

Unlike the toy spike (21 files), this simulates the regime the operator raised:
hundreds of Excel files + one ~700-"page" document. Deterministic (fixed seed).

Produces in corpus/:
  - N_INV invoice workbooks (fattura_2026_NNNN.xlsx), 40 clients, some with
    DIRTY names (versioned/duplicate) to stress filename-based routing
  - contabilita_2026.xlsx  -- ledger aggregating every invoice
  - a few report_*.xlsx     -- larger sheets
  - contratto_quadro.txt    -- the ~700-page document, sectioned, with planted
    "needle" facts in specific articles (for the flat-vs-hierarchical test)
  - nota_*.txt              -- prose notes citing specific invoices (cross edges)
  - edges_groundtruth.json  -- the true inter-document edges (for blocking test)
  - needles.json            -- planted facts + queries (for passage-recall test)
"""

import json
import os
import random
from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "corpus")

SEED = 20260801
N_INV = 400
N_CLIENTS = 40
N_NOTES = 40
BIG_DOC_SECTIONS = 140      # ~ "700 pages" of filler
PARAS_PER_SECTION = 16
VAT = 0.22

random.seed(SEED)

CLIENTS = [f"Cliente{ i:02d} {suf}" for i, suf in
           [(i, random.choice(["SpA", "SRL", "SNC", "& Figli", "Group", "Import"]))
            for i in range(1, N_CLIENTS + 1)]]


def money(x):
    return round(x, 2)


def build_invoices():
    inv = []
    for n in range(1, N_INV + 1):
        client = CLIENTS[n % N_CLIENTS]
        month = 1 + (n % 12)
        day = 1 + (n % 27)
        imp = money(500 + (n * 137 % 40000))
        iva = money(imp * VAT)
        stato = random.choice(["pagata", "pagata", "pagata", "non pagata", "scaduta"])
        inv.append({
            "id": f"2026_{n:04d}",
            "cliente": client,
            "data": f"2026-{month:02d}-{day:02d}",
            "imponibile": imp, "iva": iva, "totale": money(imp + iva),
            "stato": stato,
        })
    return inv


def write_invoice(inv, fname=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "fattura"
    ws.append(["Fattura n.", inv["id"]])
    ws.append(["Cliente", inv["cliente"]])
    ws.append(["Data emissione", inv["data"]])
    ws.append(["Stato", inv["stato"]])
    ws.append([])
    ws.append(["Descrizione", "Quantita", "Prezzo", "Importo"])
    ws.append(["Prestazione servizi", 1, inv["imponibile"], inv["imponibile"]])
    ws.append(["Imponibile", inv["imponibile"]])
    ws.append(["IVA 22%", inv["iva"]])
    ws.append(["Totale documento", inv["totale"]])
    wb.save(os.path.join(OUT, fname or f"fattura_{inv['id']}.xlsx"))


def write_ledger(inv):
    wb = Workbook()
    reg = wb.active
    reg.title = "registro"
    reg.append(["n_fattura", "cliente", "data", "imponibile", "iva", "totale", "stato"])
    for r in inv:
        reg.append([r["id"], r["cliente"], r["data"], r["imponibile"], r["iva"], r["totale"], r["stato"]])
    wb.save(os.path.join(OUT, "contabilita_2026.xlsx"))


def write_reports():
    for name, title in [("report_vendite_2026", "Report vendite annuale"),
                        ("report_iva_2026", "Riepilogo IVA trimestrale"),
                        ("report_insoluti_2026", "Report crediti insoluti")]:
        wb = Workbook()
        ws = wb.active
        ws.title = "report"
        ws.append([title])
        ws.append(["trimestre", "totale"])
        for q, v in [("Q1", 120000), ("Q2", 145000), ("Q3", 98000), ("Q4", 160000)]:
            ws.append([q, v])
        wb.save(os.path.join(OUT, f"{name}.xlsx"))


def write_notes(inv):
    edges = []
    for k in range(N_NOTES):
        target = inv[random.randrange(len(inv))]
        fname = f"nota_{k:03d}.txt"
        kind = random.choice(["contestazione", "sollecito", "chiarimento"])
        text = (f"Nota interna {kind} relativa alla fattura {target['id']} "
                f"del cliente {target['cliente']}. "
                f"Posizione da verificare con l'ufficio contabilita.\n")
        with open(os.path.join(OUT, fname), "w", encoding="utf-8") as f:
            f.write(text)
        edges.append({"src": fname, "dst": f"fattura_{target['id']}.xlsx", "type": f"nota_{kind}"})
    return edges


LOREM = ("La presente clausola disciplina i rapporti tra le parti nel rispetto "
         "della normativa vigente e degli usi commerciali applicabili al settore. "
         "Le parti si impegnano reciprocamente a garantire la corretta esecuzione "
         "degli obblighi contrattuali assunti e a comunicare tempestivamente ogni "
         "variazione rilevante. ")

# Distractor legal vocabulary sprinkled across MANY sections, so a flat BM25 over
# all chunks gets many competing matches for the query terms -- the discriminating
# signal is the section (title/summary), not the raw keyword. This makes the
# flat-vs-hierarchical comparison honest instead of trivially flat-friendly.
DISTRACTORS = [
    "Eventuali penali sono regolate secondo le disposizioni generali applicabili.",
    "La responsabilita delle parti resta soggetta ai limiti di legge.",
    "I termini indicati decorrono dalla data di sottoscrizione.",
    "Le comunicazioni in materia di riservatezza seguono la procedura ordinaria.",
    "Per quanto non previsto si rinvia al foro e alla normativa generale.",
    "Eventuali interessi seguono le condizioni economiche pattuite.",
    "Il recesso ordinario e' disciplinato dalle disposizioni comuni.",
    "Il subappalto resta soggetto alle regole generali del presente contratto.",
]

# planted needles: (section_number, section_title, needle_sentence, query)
NEEDLE_SPECS = [
    (12, "Recesso anticipato", "Il recesso anticipato comporta una penale del 15 per cento del residuo.",
     "quale penale si applica in caso di recesso anticipato"),
    (23, "Riservatezza", "Gli obblighi di riservatezza permangono per cinque anni dopo la cessazione.",
     "per quanto tempo valgono gli obblighi di riservatezza dopo la fine del contratto"),
    (41, "Foro competente", "Per ogni controversia e' competente in via esclusiva il foro di Milano.",
     "quale foro e' competente per le controversie"),
    (58, "Limitazione di responsabilita", "La responsabilita e' limitata al valore di dodici mensilita di corrispettivo.",
     "a quanto e' limitata la responsabilita del fornitore"),
    (77, "Subappalto", "Il subappalto e' ammesso solo previa autorizzazione scritta del committente.",
     "e' possibile subappaltare le prestazioni"),
    (94, "Forza maggiore", "Gli eventi di forza maggiore sospendono i termini per un massimo di novanta giorni.",
     "per quanto tempo la forza maggiore sospende i termini"),
    (108, "Proprieta intellettuale", "Tutti i diritti di proprieta intellettuale restano in capo al committente.",
     "a chi appartengono i diritti di proprieta intellettuale"),
    (129, "Interessi di mora", "In caso di ritardo si applicano interessi di mora al tasso legale maggiorato di otto punti.",
     "quali interessi di mora si applicano in caso di ritardo di pagamento"),
]


def write_big_doc():
    needles = {s[0]: s for s in NEEDLE_SPECS}
    lines = ["CONTRATTO QUADRO DI FORNITURA - documento generale\n"]
    meta_needles = []
    for sec in range(1, BIG_DOC_SECTIONS + 1):
        if sec in needles:
            _, title, needle, query = needles[sec]
        else:
            title = f"Disposizioni generali comma {sec}"
            needle = None
        lines.append(f"\n## Articolo {sec} - {title}\n")
        for p in range(PARAS_PER_SECTION):
            lines.append(LOREM * 2 + random.choice(DISTRACTORS) + " ")
            if needle and p == PARAS_PER_SECTION // 2:
                lines.append(needle + " ")
        if sec in needles:
            _, title, needle, query = needles[sec]
            meta_needles.append({"section": sec, "title": title,
                                 "needle": needle, "query": query})
    with open(os.path.join(OUT, "contratto_quadro.txt"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return meta_needles


def main():
    os.makedirs(OUT, exist_ok=True)
    inv = build_invoices()
    for r in inv:
        write_invoice(r)
    # dirty names: 20 versioned duplicates of real invoices
    for r in inv[:20]:
        suffix = random.choice(["_v2", "_FINALE", " (copia)", "_rev"])
        write_invoice(r, fname=f"fattura_{r['id']}{suffix}.xlsx")
    write_ledger(inv)
    write_reports()
    note_edges = write_notes(inv)
    needles = write_big_doc()

    # ground-truth edges: same-client (a sample) + note->invoice
    client_edges = []
    by_client = {}
    for r in inv:
        by_client.setdefault(r["cliente"], []).append(f"fattura_{r['id']}.xlsx")
    for c, files in by_client.items():
        for i in range(len(files) - 1):
            client_edges.append({"src": files[i], "dst": files[i + 1],
                                 "type": "stesso_cliente", "key": c})
    with open(os.path.join(OUT, "edges_groundtruth.json"), "w") as f:
        json.dump({"client_edges": client_edges, "note_edges": note_edges,
                   "by_client": by_client}, f)
    with open(os.path.join(OUT, "needles.json"), "w") as f:
        json.dump(needles, f)

    files = os.listdir(OUT)
    big = os.path.getsize(os.path.join(OUT, "contratto_quadro.txt"))
    print(f"generated {len(files)} files")
    print(f"  invoices ~{N_INV} (+20 dirty-named), {N_CLIENTS} clients, {N_NOTES} notes")
    print(f"  big doc: {BIG_DOC_SECTIONS} sections, {big/1024:.0f} KB "
          f"(~{big//2000} 'pages'), {len(needles)} planted needles")


if __name__ == "__main__":
    main()
